# --------------------------------------------------------
# Ke Yan,
# Imaging Biomarkers and Computer-Aided Diagnosis Laboratory (CADLab)
# National Institutes of Health Clinical Center,
# Apr 2019.
# This file contains codes to load and save information from/to files.
# --------------------------------------------------------

from openpyxl import load_workbook, Workbook
from config import config, default
import numpy as np
import os
import csv
from scipy.io import savemat


cellname = lambda row, col: '%s%d' % (chr(ord('A') + col - 1), row)


def save_to_xlsx_2d(fn, col_names, row_names, mat):
    wb = Workbook()
    sheet = wb.active
    for i, key in enumerate(row_names):
        sheet[cellname(i + 2, 1)] = key
    for i, key in enumerate(col_names):
        sheet[cellname(1, i + 2)] = key

    for row in range(len(row_names)):
        for col in range(len(col_names)):
            sheet[cellname(row + 2, col+2)] = mat[row, col]
    res = wb.save(fn)
    return res


def save_acc_to_file(acc_all, val_loader, prefix):
    rownames = default.term_list
    colnames = ['train_positive', 'test_positive', 'test_negative', 'auc', 'precision', 'recall', 'f1', 'threshold']
    matdata = np.vstack((default.cls_sz_train, val_loader.dataset.cls_sz, acc_all['ex_neg'], acc_all['auc_perclass'],
                         acc_all['perclass_precisions'], acc_all['perclass_recalls'], acc_all['perclass_f1s'],
                         default.best_cls_ths))
    filename = os.path.join(default.res_path, '%s_%s.xlsx' % (prefix, default.exp_name))
    if not os.path.exists(default.res_path):
        os.mkdir(default.res_path)
    save_to_xlsx_2d(filename, colnames, rownames, matdata.transpose())


def save_test_scores_to_file(score_all, pred_label_all, target_all, accs, lesion_idxs):
    fnout = default.res_path + default.exp_name + '_output.mat'
    data = {'scores': score_all, 'pred_labels': pred_label_all, 'gt_labels': target_all,
            'aucs': accs['auc_perclass'], 'term_list': default.term_list,
            'lesion_idx': lesion_idxs}
    savemat(fnout, data)


def save_ft_to_file(embedding):
    fnout = default.res_path + default.exp_name + '_embedding.mat'
    data = {'embedding': embedding}
    savemat(fnout, data)


def load_DL_info(path):
    # load annotations and meta-info from DL_info.csv
    info = []
    with open(path, 'rb') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            filename = row[0]  # replace the last _ in filename with / or \
            idx = filename.rindex('_')
            row[0] = filename[:idx] + os.sep + filename[idx+1:]
            info.append(row)
    info = info[1:]

    # the information not used in this project are commented
    res = {}
    res['filenames'] = np.array([row[0] for row in info])
    res['patient_idx'] = np.array([int(row[1]) for row in info])
    # info['study_idx'] = np.array([int(row[2]) for row in info])
    # info['series_idx'] = np.array([int(row[3]) for row in info])
    res['slice_idx'] = np.array([int(row[4]) for row in info])
    # res['d_coordinate'] = np.array([[float(x) for x in row[5].split(',')] for row in info])
    # res['d_coordinate'] -= 1
    res['boxes'] = np.array([[float(x) for x in row[6].split(',')] for row in info])
    res['boxes'] -= 1  # coordinates in info file start from 1
    # res['diameter'] = np.array([[float(x) for x in row[7].split(',')] for row in info])
    res['norm_location'] = np.array([[float(x) for x in row[8].split(',')] for row in info])
    res['type'] = np.array([int(row[9]) for row in info])
    res['noisy'] = np.array([int(row[10]) > 0 for row in info])
    # res['slice_range'] = np.array([[int(x) for x in row[11].split(',')] for row in info])
    res['spacing3D'] = np.array([[float(x) for x in row[12].split(',')] for row in info])
    res['spacing'] = res['spacing3D'][:, 0]
    res['slice_intv'] = res['spacing3D'][:, 2]  # slice intervals
    # res['image_size'] = np.array([[int(x) for x in row[13].split(',')] for row in info])
    # res['DICOM_window'] = np.array([[float(x) for x in row[14].split(',')] for row in info])
    # res['gender'] = np.array([row[15] for row in info])
    # res['age'] = np.array([float(row[16]) for row in info])  # may be NaN
    res['train_val_test'] = np.array([int(row[17]) for row in info])

    return res


def load_ontology_from_xlsfile(fn):
    wb = load_workbook(fn)
    sheet = wb.active
    term_dicts = []
    for p in range(2, sheet.max_row + 1):
        parents = sheet[cellname(p, 7)].value
        parents = [] if parents is None else parents.split(' | ')
        children = sheet[cellname(p, 8)].value
        children = [] if children is None else children.split(' | ')
        ex = sheet[cellname(p, 6)].value
        ex = [] if ex is None else ex.split(' | ')
        term_dict = {'id': sheet[cellname(p, 1)].value,
                     'class': sheet[cellname(p, 2)].value,
                     'term': sheet[cellname(p, 3)].value,
                     'synonyms': sheet[cellname(p, 4)].value.split(' | '),
                     'num_detected': sheet[cellname(p, 5)].value,
                     'exclusive': ex,
                     'parents': parents,
                     'children': children}
        term_dicts.append(term_dict)
    return term_dicts


def load_demo_labels(fn):
    wb = load_workbook(fn)
    sheet = wb.active
    terms = []
    thresolds = []
    for p in range(2, sheet.max_row+1):
        terms.append(sheet[cellname(p, 1)].value)
        thresolds.append(float(sheet[cellname(p, 9)].value))
    return terms, thresolds
