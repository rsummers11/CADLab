# Ke Yan, Imaging Biomarkers and Computer-Aided Diagnosis Laboratory,
# National Institutes of Health Clinical Center, July 2019
"""Utilities for DeepLesion"""
import numpy as np
from openpyxl import load_workbook
import json
from collections import Counter
from maskrcnn.utils.miscellaneous import unique
from maskrcnn.config import cfg


def gen_mask_polygon_from_recist(recist):
    """Generate ellipse from RECIST for weakly-supervised segmentation"""
    x11, y11, x12, y12, x21, y21, x22, y22 = recist
    axis1 = np.linalg.solve(np.array([[x11, y11], [x12, y12]]), np.array([1, 1]))
    axis2 = np.linalg.solve(np.array([[x21, y21], [x22, y22]]), np.array([1, 1]))
    center = np.linalg.solve(np.array([[axis1[0], axis1[1]], [axis2[0], axis2[1]]]), np.array([1, 1]))
    centered_recist = recist - np.tile(center, (4,))
    centered_recist = np.reshape(centered_recist, (4, 2))
    pt_angles = np.arctan2(centered_recist[:, 1], centered_recist[:, 0])
    pt_lens = np.sqrt(np.sum(centered_recist ** 2, axis=1))

    ord = [0, 2, 1, 3, 0]
    grid = .1
    rotated_pts = []
    for p in range(4):
        # pt1 = centered_recist[ord[p]]
        # pt2 = centered_recist[ord[p+1]]
        if (pt_angles[ord[p]] < pt_angles[ord[p + 1]] and pt_angles[ord[p + 1]] - pt_angles[ord[p]] < np.pi) \
                or (pt_angles[ord[p]] - pt_angles[ord[p + 1]] > np.pi):  # counter-clockwise
            angles = np.arange(0, np.pi / 2, grid)
        else:
            angles = np.arange(0, -np.pi / 2, -grid)

        xs = np.cos(angles) * pt_lens[ord[p]]
        ys = np.sin(angles) * pt_lens[ord[p + 1]]
        r = pt_angles[ord[p]]
        rotated_pts1 = np.matmul(np.array([[np.cos(r), -np.sin(r)], [np.sin(r), np.cos(r)]]),
                                 np.vstack((xs, ys)))
        rotated_pts.append(rotated_pts1)
    rotated_pts = np.hstack(rotated_pts)
    decentered_pts = rotated_pts + center.reshape((2, 1))
    polygon = decentered_pts.transpose().ravel()
    # for p in polygon:
    #     print('%.4f'%p, ',',)
    # print('\n',recist)
    return polygon.tolist()


def load_tag_dict_from_xlsfile(fn):
    """Load ontology"""
    cellname = lambda row, col: '%s%d' % (chr(ord('A') + col - 1), row)
    wb = load_workbook(fn)
    sheet = wb.get_active_sheet()
    tag_dicts = []
    for p in range(2, sheet.max_row + 1):
        ex = sheet[cellname(p, 6)].value
        ex = [] if ex is None else ex.split(' | ')
        parents = sheet[cellname(p, 7)].value
        parents = [] if parents is None else parents.split(' | ')
        children = sheet[cellname(p, 8)].value
        children = [] if children is None else children.split(' | ')
        tag_dict = {'id': sheet[cellname(p, 1)].value,  # useless
                     'class': sheet[cellname(p, 2)].value,
                     'tag': sheet[cellname(p, 3)].value,
                     'synonyms': sheet[cellname(p, 4)].value.split(' | '),
                     'num_detected': sheet[cellname(p, 5)].value,
                     'exclusive': ex,
                     'parents': parents,
                     'children': children
                    }
        tag_dicts.append(tag_dict)
    return tag_dicts


def load_lesion_tags(split_file, tag_dict):
    """Load training labels for tags"""
    with open(split_file, 'r') as f:
        data = json.load(f)
        print('loaded', split_file)
    term_list = data['term_list']
    num_labels = len(term_list)

    prefix = 'train'
    smp_idxs, labels, uncertain_labels = \
        data['%s_lesion_idxs' % prefix], data['%s_relevant_labels' % prefix], \
        data['%s_uncertain_labels' % prefix]

    tag_dict_filtered = {idx: unique(r+u) for idx,r,u in zip(smp_idxs, labels, uncertain_labels)}
    tag_list_dict = []
    class_map = {t['tag']: t['class'] for t in tag_dict}
    for i in range(num_labels):
        tag_dict = {'ID': i, 'tag': term_list[i], 'class': class_map[term_list[i]]}
        tag_list_dict.append(tag_dict)
    return tag_list_dict, tag_dict_filtered


def gen_parent_list(tag_dicts, tag_list):
    """Hierarchical label relations"""
    parents_map = {t['tag']: t['parents'] for t in tag_dicts}
    parent_list = []
    for t in tag_list:
        ps = parents_map[t]
        parent_list.append([tag_list.index(p) for p in ps if p in tag_list])
    return parent_list


def gen_children_list(parent_list, tag_list):
    """Hierarchical label relations"""
    all_children_list = [[] for _ in tag_list]
    for i, parent in enumerate(parent_list):
        for p1 in parent:
            all_children_list[p1].append(i)

    direct_children_list = [[] for _ in tag_list]
    for i, children in enumerate(all_children_list):
        direct_children_list[i] = [c for c in children if not any([p in children for p in parent_list[c]])]
    return all_children_list, direct_children_list


def gen_tree_depth(tag_list, parent_list):
    """Hierarchical label relations"""
    tag_depth = np.ones((len(tag_list),), dtype=int)
    while True:
        last_depth = tag_depth.copy()
        for p in range(len(parent_list)):
            if len(parent_list[p]) > 0:
                tag_depth[p] = np.max([tag_depth[idx] for idx in parent_list[p]])+1
        if np.all(last_depth == tag_depth):
            break
    return tag_depth


def gen_exclusive_list(tag_dicts, tag_list, parent_list, all_children_list):
    """Infer exclusive label relations according to hierarchical relations"""
    exclusive_list = []
    all_d_tags = [t['tag'] for t in tag_dicts]
    for p in range(len(tag_list)):
        idx = all_d_tags.index(tag_list[p])
        exclusive_list.append([tag_list.index(ex) for ex in
                                    tag_dicts[idx]['exclusive'] if ex in tag_list])
    while True:
        flag = False
        for p in range(len(tag_list)):
            cur_ex = exclusive_list[p]
            next_ex = cur_ex[:]
            for ex in cur_ex:
                next_ex += all_children_list[ex]
            for parent in parent_list[p]:
                next_ex += exclusive_list[parent]
            next_ex = unique(next_ex)
            flag = flag or (set(next_ex) != set(cur_ex))
            exclusive_list[p] = next_ex
        if not flag:
            break

    return exclusive_list
